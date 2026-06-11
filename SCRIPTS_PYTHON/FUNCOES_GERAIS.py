# =========================================================
# PROJETO : AUTOMACAO SINAN
# ARQUIVO : FUNCOES_GERAIS.py
# OBJETIVO:
# Motor central operacional epidemiologico.
# =========================================================

# =========================================================
# IMPORTACOES
# =========================================================

from pathlib import Path

from datetime import datetime, date, timedelta

import pandas as pd

from dbfread import DBF

# =========================================================
# IMPORTAR CONFIG
# =========================================================

from CONFIG_00 import (

    AGRAVO_ATIVO,

    MAPA_AGRAVOS,

    MUNICIPIO_PADRAO,

    ANO_EPIDEMIOLOGICO,

    PASTA_LOGS,

    VALORES_INVALIDOS,

    COLUNAS_MUNICIPIO,

    STATUS_CONFIRMADO,

    STATUS_DESCARTADO,

    STATUS_INCONCLUSIVO

)

# =========================================================
# ESCREVER LOG
# =========================================================

def escrever_log(

    mensagem,

    tipo="EXECUCAO"

):

    try:

        data_hoje = datetime.now().strftime(
            "%Y_%m_%d"
        )

        pasta_log = (

            PASTA_LOGS /
            tipo.upper()

        )

        pasta_log.mkdir(

            parents=True,

            exist_ok=True

        )

        arquivo_log = (

            pasta_log /
            f"log_{data_hoje}.txt"

        )

        horario = datetime.now().strftime(
            "%H:%M:%S"
        )

        with open(

            arquivo_log,

            "a",

            encoding="utf-8"

        ) as log:

            log.write(

                f"[{horario}] {mensagem}\n"

            )

    except Exception as erro:

        print(

            f"[ERRO LOG] {erro}"

        )

# =========================================================
# LER DBF
# =========================================================

def ler_dbf(caminho_dbf):

    try:

        tabela = DBF(

            caminho_dbf,

            encoding="latin1",

            ignore_missing_memofile=True,

            char_decode_errors="ignore"

        )

        df = pd.DataFrame(

            iter(tabela)

        )

        escrever_log(

            f"DBF carregado: "
            f"{Path(caminho_dbf).name}"

        )

        return df

    except Exception as erro:

        escrever_log(

            f"Erro leitura DBF: {erro}",

            tipo="ERROS"

        )

        raise

# =========================================================
# LOCALIZAR DBF
# =========================================================

def localizar_dbf(agravo=None):

    if agravo is None:

        agravo = AGRAVO_ATIVO

    if agravo not in MAPA_AGRAVOS:

        raise FileNotFoundError(

            f"Agravo nao configurado: {agravo}"

        )

    pasta_dbf = (

        MAPA_AGRAVOS[
            agravo
        ]["pasta"]

    )

    arquivos_dbf = (

        list(pasta_dbf.glob("*.dbf"))

        +

        list(pasta_dbf.glob("*.DBF"))

    )

    if len(arquivos_dbf) == 0:

        raise FileNotFoundError(

            f"Nenhum DBF encontrado para o agravo "
            f"'{agravo}' em: {pasta_dbf}"

        )

    arquivo_mais_recente = max(

        arquivos_dbf,

        key=lambda caminho: caminho.stat().st_mtime

    )

    return arquivo_mais_recente

# =========================================================
# PADRONIZAR TEXTO
# =========================================================

def padronizar_texto(serie):

    return (

        serie
        .fillna("")
        .astype(str)
        .str.strip()

    )

# =========================================================
# DETECTAR TIPO EPIDEMIOLOGICO
# =========================================================

def detectar_tipo_coluna(serie):

    try:

        amostra = (

            serie
            .dropna()
            .astype(str)
            .head(50)

        )

        if len(amostra) == 0:

            return "texto"

        # =================================================
        # DATAS
        # =================================================

        total_data = 0

        for valor in amostra:

            valor = str(valor).strip()

            if (

                "-" in valor
                or
                "/" in valor

            ):

                total_data += 1

        if total_data >= (

            len(amostra) * 0.7

        ):

            return "data"

        # =================================================
        # NUMERICO
        # =================================================

        total_num = 0

        for valor in amostra:

            valor = (

                str(valor)
                .replace(".0", "")
                .strip()

            )

            if valor.isdigit():

                total_num += 1

        if total_num >= (

            len(amostra) * 0.8

        ):

            return "numerico"

        return "texto"

    except Exception:

        return "texto"

# =========================================================
# CONVERTER COLUNAS EPIDEMIOLOGICAS
# =========================================================

def converter_colunas_epidemiologicas(df):

    df = df.copy()

    for coluna in df.columns:

        try:

            tipo_detectado = detectar_tipo_coluna(

                df[coluna]

            )

            # =============================================
            # DATA
            # =============================================

            if (

                coluna.startswith("DT_")
                or
                tipo_detectado == "data"

            ):

                serie_original = (

                    df[coluna]
                    .copy()

                )

                serie = (

                    serie_original
                    .astype(str)
                    .str.strip()
                )

                serie = serie.replace({

                    "": pd.NA,
                    "nan": pd.NA,
                    "None": pd.NA,
                    "NaT": pd.NA,
                    "00000000": pd.NA

                })

                convertido = pd.to_datetime(

                    serie,

                    errors="coerce",

                    format="mixed"

                )

                total_validos = (

                    convertido.notna().sum()

                )

                total_original = (

                    serie.notna().sum()

                )

                # =========================================
                # SOMENTE CONVERTER SE FOR SEGURO
                # =========================================

                if total_original > 0:

                    percentual = (

                        total_validos /
                        total_original

                    )

                else:

                    percentual = 0

                if percentual >= 0.60:

                    df[coluna] = convertido

                    escrever_log(

                        f"{coluna}: "
                        f"convertido DATA "
                        f"({round(percentual*100,2)}%)"

                    )

                else:

                    df[coluna] = serie_original

                    escrever_log(

                        f"{coluna}: "
                        f"mantido ORIGINAL "
                        f"(baixa confianca)",

                        tipo="ALERTAS"

                    )

            # =============================================
            # NUMERICO
            # =============================================

            elif tipo_detectado == "numerico":

                df[coluna] = (

                    df[coluna]
                    .fillna("")
                    .astype(str)
                    .str.replace(".0", "", regex=False)
                    .str.strip()

                )

            # =============================================
            # TEXTO
            # =============================================

            else:

                df[coluna] = (

                    df[coluna]
                    .fillna("")
                    .astype(str)
                    .str.strip()

                )

        except Exception as erro:

            escrever_log(

                f"Erro coluna {coluna}: {erro}",

                tipo="ALERTAS"

            )

    return df

# =========================================================
# FILTRAR MUNICIPIO
# =========================================================

def filtrar_municipio(

    df,

    codigo_municipio=MUNICIPIO_PADRAO

):

    df = df.copy()

    coluna_municipio = None

    for coluna in COLUNAS_MUNICIPIO:

        if coluna in df.columns:

            coluna_municipio = coluna

            break

    if coluna_municipio is None:

        escrever_log(

            "Coluna municipio inexistente.",

            tipo="ERROS"

        )

        return df

    total_antes = len(df)

    df[coluna_municipio] = (

        df[coluna_municipio]
        .astype(str)
        .str.extract(r"(\d+)")[0]
        .str.zfill(6)

    )

    codigo_municipio = (

        str(codigo_municipio)
        .zfill(6)

    )

    df = df[

        df[coluna_municipio]
        == codigo_municipio

    ].copy()

    total_depois = len(df)

    escrever_log(

        f"Filtro municipio: "
        f"{total_antes} -> "
        f"{total_depois}"

    )

    return df

# =========================================================
# EXPORTAR EXCEL
# =========================================================

def exportar_excel(

    caminho_saida,

    dados

):

    from openpyxl import load_workbook

    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Alignment
    from openpyxl.styles import Border
    from openpyxl.styles import Side

    from openpyxl.utils import get_column_letter

    caminho_saida = Path(caminho_saida)

    caminho_saida.parent.mkdir(

        parents=True,

        exist_ok=True

    )

    try:

        # =================================================
        # EXPORTAR
        # =================================================

        if isinstance(dados, dict):

            with pd.ExcelWriter(

                caminho_saida,

                engine="openpyxl"

            ) as writer:

                for nome_aba, df in dados.items():

                    if df is None:

                        continue

                    if not isinstance(

                        df,

                        pd.DataFrame

                    ):

                        continue

                    nome_aba = (

                        str(nome_aba)[:31]

                    )

                    df.to_excel(

                        writer,

                        sheet_name=nome_aba,

                        index=False

                    )

        elif isinstance(

            dados,

            pd.DataFrame

        ):

            dados.to_excel(

                caminho_saida,

                index=False

            )

        else:

            raise Exception(

                "Formato invalido."

            )

        # =================================================
        # FORMATAR EXCEL
        # =================================================

        wb = load_workbook(

            caminho_saida

        )

        fonte_cabecalho = Font(

            bold=True,

            color="FFFFFF"

        )

        preenchimento = PatternFill(

            start_color="1F4E78",

            end_color="1F4E78",

            fill_type="solid"

        )

        alinhamento = Alignment(

            horizontal="center",

            vertical="center",

            wrap_text=True

        )

        borda = Border(

            left=Side(style="thin"),

            right=Side(style="thin"),

            top=Side(style="thin"),

            bottom=Side(style="thin")

        )

        for ws in wb.worksheets:

            ws.freeze_panes = "A2"

            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:

                cell.font = fonte_cabecalho
                cell.fill = preenchimento
                cell.alignment = alinhamento
                cell.border = borda

            for coluna in ws.columns:

                tamanho_maximo = 0

                letra_coluna = get_column_letter(

                    coluna[0].column

                )

                for cell in coluna[:100]:

                    try:

                        valor = str(cell.value)

                        tamanho = len(valor)

                        if tamanho > tamanho_maximo:

                            tamanho_maximo = tamanho

                    except Exception:

                        pass

                largura = min(

                    tamanho_maximo + 3,

                    60

                )

                ws.column_dimensions[
                    letra_coluna
                ].width = largura

        wb.save(

            caminho_saida

        )

        escrever_log(

            f"Excel exportado: "
            f"{caminho_saida.name}"

        )

        print(

            f"[OK] Excel exportado: "
            f"{caminho_saida.name}"

        )

    except Exception as erro:

        escrever_log(

            f"Erro exportacao: {erro}",

            tipo="ERROS"

        )

        raise

# =========================================================
# CALCULAR PERCENTUAL
# =========================================================

def calcular_percentual(total, base):

    if base == 0:

        return 0

    return round(

        (total / base) * 100,

        2

    )

# =========================================================
# GARANTIR DATETIME
# =========================================================

def garantir_datetime(df, colunas):

    for coluna in colunas:

        if coluna not in df.columns:

            escrever_log(

                f"Coluna de data ausente: {coluna}",

                tipo="AVISOS"

            )

            continue

        try:

            df[coluna] = pd.to_datetime(

                df[coluna],

                errors="coerce"

            )

        except Exception as erro:

            escrever_log(

                f"Falha ao converter {coluna} para datetime: {erro}",

                tipo="AVISOS"

            )

            continue

        if df[coluna].isna().all():

            escrever_log(

                f"Coluna {coluna} sem datas validas apos conversao.",

                tipo="AVISOS"

            )

    return df

# =========================================================
# FILTRAR CONFIRMADOS
# =========================================================

def filtrar_confirmados(df):

    if "STATUS_MS" not in df.columns:

        return df

    return df[

        df["STATUS_MS"]
        == STATUS_CONFIRMADO

    ].copy()

# =========================================================
# FILTRAR DESCARTADOS
# =========================================================

def filtrar_descartados(df):

    if "STATUS_MS" not in df.columns:

        return df

    return df[

        df["STATUS_MS"]
        == STATUS_DESCARTADO

    ].copy()

# =========================================================
# FILTRAR PROVAVEIS
# =========================================================

def filtrar_provaveis(df):

    if "STATUS_MS" not in df.columns:

        return df

    return df[

        df["STATUS_MS"]
        != STATUS_DESCARTADO

    ].copy()

# =========================================================
# FILTRAR INDICADORES
# =========================================================

def filtrar_indicadores(df):

    if "USAR_INDICADOR" not in df.columns:

        return df

    return df[

        df["USAR_INDICADOR"]
        == "SIM"

    ].copy()

# =========================================================
# RESUMO RAPIDO
# =========================================================

def resumo_rapido(df):

    resumo = {

        "TOTAL_REGISTROS": len(df),

        "TOTAL_COLUNAS": len(df.columns)

    }

    if "STATUS_MS" in df.columns:

        resumo["CONFIRMADOS"] = (

            df["STATUS_MS"]
            .eq(STATUS_CONFIRMADO)
            .sum()

        )

        resumo["DESCARTADOS"] = (

            df["STATUS_MS"]
            .eq(STATUS_DESCARTADO)
            .sum()

        )

        resumo["INCONCLUSIVOS"] = (

            df["STATUS_MS"]
            .eq(STATUS_INCONCLUSIVO)
            .sum()

        )

    escrever_log(

        f"Resumo: {resumo}"

    )

    return resumo

# =========================================================
# PREPARAR BASE
# =========================================================

def calcular_inicio_ano_epidemiologico(ano):

    # =====================================================
    # REGRA MS:
    # - a semana epidemiologica comeca no domingo
    # - a SE 1 e a semana que contem a maior parte
    #   (4 ou mais) dos primeiros 7 dias de janeiro
    # =====================================================

    primeiro_janeiro = date(ano, 1, 1)

    # dias decorridos desde o domingo anterior
    # (weekday: segunda=0 ... domingo=6) -> domingo=0

    dias_desde_domingo = (
        primeiro_janeiro.weekday() + 1
    ) % 7

    domingo_da_semana = (
        primeiro_janeiro
        - timedelta(days=dias_desde_domingo)
    )

    # =====================================================
    # dias dessa semana que caem no ano novo:
    # 7 - dias_desde_domingo. Se >= 4 (maioria),
    # a SE 1 e a semana que contem 1 de janeiro.
    # Caso contrario, a SE 1 comeca no domingo seguinte.
    # =====================================================

    if (7 - dias_desde_domingo) >= 4:

        inicio_se1 = domingo_da_semana

    else:

        inicio_se1 = (
            domingo_da_semana
            + timedelta(days=7)
        )

    return pd.Timestamp(inicio_se1)


def preparar_base(

    df,

    aplicar_filtro_municipio=True

):

    print("\n===================================")
    print(" PREPARANDO BASE EPIDEMIOLOGICA ")
    print("===================================\n")

    # =====================================================
    # TIPAGEM EPIDEMIOLOGICA
    # =====================================================

    df = converter_colunas_epidemiologicas(df)

    # =====================================================
    # FILTRO MUNICIPIO
    # =====================================================

    if aplicar_filtro_municipio:

        df = filtrar_municipio(df)

    # =====================================================
    # FILTRO ANO EPIDEMIOLOGICO
    # =====================================================

    if "DT_SIN_PRI" in df.columns:

        # Corte em 1 de janeiro do ano epidemiologico.
        # (calcular_inicio_ano_epidemiologico fica disponivel
        #  para relatorios por SE, mas aqui o corte e 01/01.)
        data_corte = pd.Timestamp(
            ANO_EPIDEMIOLOGICO, 1, 1
        )

        total_antes = len(df)

        df = df[
            df["DT_SIN_PRI"] >= data_corte
        ].copy()

        total_depois = len(df)

        escrever_log(
            f"Filtro ano epidemiologico: "
            f"corte {data_corte.date()} | "
            f"{total_antes} -> {total_depois}"
        )

    # =====================================================
    # TRADUCOES
    # =====================================================

    try:

        from FUNCOES_TRADUCAO import (

            aplicar_traducoes,

            obter_status_ms,

            obter_usar_indicador

        )

        df = aplicar_traducoes(df)

        df = obter_status_ms(df)

        df = obter_usar_indicador(df)

    except Exception as erro:

        escrever_log(

            f"Erro traducoes: {erro}",

            tipo="ERROS"

        )

    resumo = resumo_rapido(df)

    print(resumo)

    print("\n===================================")
    print(" BASE PREPARADA ")
    print("===================================\n")

    return df